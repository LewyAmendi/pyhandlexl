"""Tests for the internal _safety module."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import pyhandlexl._safety as safety
from pyhandlexl._safety import atomic_save, load_or_create, safe_load
from pyhandlexl.errors import FileLockedError, InvalidFileError


@pytest.fixture(autouse=True)
def _no_sleep(monkeypatch):
    """Make retries instant so the suite stays fast."""
    monkeypatch.setattr(safety, "sleep", lambda _seconds: None)


def _make_workbook(path: Path, a1: str = "hello") -> None:
    wb = Workbook()
    wb.active["A1"] = a1
    wb.save(path)


class TestRetry:
    def test_succeeds_after_transient_failures(self):
        attempts = []

        def flaky():
            attempts.append(1)
            if len(attempts) < 3:
                raise PermissionError("locked")
            return "done"

        assert safety._retry(flaky, path=Path("x.xlsx"), retries=5) == "done"
        assert len(attempts) == 3

    def test_raises_file_locked_when_exhausted(self):
        def always_locked():
            raise PermissionError("still locked")

        with pytest.raises(FileLockedError):
            safety._retry(always_locked, path=Path("x.xlsx"), retries=3)

    def test_non_transient_error_is_not_retried(self):
        attempts = []

        def boom():
            attempts.append(1)
            raise ValueError("not a lock")

        with pytest.raises(ValueError):
            safety._retry(boom, path=Path("x.xlsx"), retries=5)
        assert len(attempts) == 1


class TestSafeLoad:
    def test_loads_existing_workbook(self, tmp_path):
        p = tmp_path / "wb.xlsx"
        _make_workbook(p)
        assert safe_load(p).active["A1"].value == "hello"

    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            safe_load(tmp_path / "nope.xlsx")

    def test_invalid_file_raises(self, tmp_path):
        p = tmp_path / "bad.xlsx"
        p.write_text("not a workbook")
        with pytest.raises(InvalidFileError):
            safe_load(p)


class TestLoadOrCreate:
    def test_returns_new_workbook_when_absent(self, tmp_path):
        assert isinstance(load_or_create(tmp_path / "fresh.xlsx"), Workbook)

    def test_loads_when_present(self, tmp_path):
        p = tmp_path / "wb.xlsx"
        _make_workbook(p, a1="existing")
        assert load_or_create(p).active["A1"].value == "existing"


class TestAtomicSave:
    def test_creates_new_file(self, tmp_path):
        p = tmp_path / "new.xlsx"
        atomic_save(Workbook(), p)
        assert p.is_file()

    def test_overwrites_existing_content(self, tmp_path):
        p = tmp_path / "wb.xlsx"
        _make_workbook(p, a1="old")
        wb = Workbook()
        wb.active["A1"] = "new"
        atomic_save(wb, p)
        assert load_workbook(p).active["A1"].value == "new"

    def test_no_temp_file_left_on_success(self, tmp_path):
        p = tmp_path / "wb.xlsx"
        atomic_save(Workbook(), p)
        assert {f.name for f in tmp_path.iterdir()} == {"wb.xlsx"}

    def test_original_preserved_when_validation_fails(self, tmp_path, monkeypatch):
        p = tmp_path / "wb.xlsx"
        _make_workbook(p, a1="keep me")

        monkeypatch.setattr(safety, "is_valid_xlsx", lambda _path: False)
        doomed = Workbook()
        doomed.active["A1"] = "should not land"
        with pytest.raises(InvalidFileError):
            atomic_save(doomed, p)

        assert load_workbook(p).active["A1"].value == "keep me"
        assert {f.name for f in tmp_path.iterdir()} == {"wb.xlsx"}

    def test_locked_target_raises_file_locked_and_cleans_up(self, tmp_path, monkeypatch):
        p = tmp_path / "wb.xlsx"
        _make_workbook(p)

        def locked_replace(src, dst):
            raise PermissionError("target open in Excel")

        monkeypatch.setattr(safety.os, "replace", locked_replace)
        with pytest.raises(FileLockedError):
            atomic_save(Workbook(), p)

        assert {f.name for f in tmp_path.iterdir()} == {"wb.xlsx"}
