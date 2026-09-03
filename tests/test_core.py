"""Tests for core.py: read_sheet, write_sheet, append_rows."""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from pyhandlexl.core import (
    append_rows,
    create_sheet,
    delete_sheet,
    list_sheets,
    read_sheet,
    rename_sheet,
    sheet_exists,
    write_sheet,
)
from pyhandlexl.errors import SheetNameError, SheetNotFoundError


class TestWriteThenRead:
    def test_round_trip_on_new_file(self, tmp_path):
        path = tmp_path / "out.xlsx"
        write_sheet(path, [["a", "b"], ["c", "d"]])
        assert read_sheet(path) == [["a", "b"], ["c", "d"]]

    def test_new_file_has_only_the_written_sheet(self, tmp_path):
        path = tmp_path / "out.xlsx"
        write_sheet(path, [["x"]], sheet="Data")
        assert load_workbook(path).sheetnames == ["Data"]

    def test_values_are_returned_as_strings(self, tmp_path):
        path = tmp_path / "out.xlsx"
        write_sheet(path, [[1, 2.5, True, None]])
        assert read_sheet(path) == [["1", "2.5", "True"]]  # trailing None trimmed

    def test_write_replaces_existing_content(self, tmp_path):
        path = tmp_path / "out.xlsx"
        write_sheet(path, [["old", "old", "old"], ["old", "old", "old"]])
        write_sheet(path, [["new"]])
        assert read_sheet(path) == [["new"]]

    def test_other_sheets_are_preserved(self, tmp_path):
        path = tmp_path / "book.xlsx"
        wb = Workbook()
        wb.active.title = "Keep"
        wb.active["A1"] = "untouched"
        wb.create_sheet("Target")
        wb.save(path)

        write_sheet(path, [["changed"]], sheet="Target")

        result = load_workbook(path)
        assert result["Keep"]["A1"].value == "untouched"
        assert read_sheet(path, "Target") == [["changed"]]


class TestReadSheet:
    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            read_sheet(tmp_path / "nope.xlsx")

    def test_unknown_sheet_raises(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]])
        with pytest.raises(SheetNotFoundError):
            read_sheet(path, "Ghost")

    def test_trailing_empty_cells_are_trimmed(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a", "", ""], ["b", "c", ""]])
        assert read_sheet(path) == [["a"], ["b", "c"]]

    def test_pad_makes_result_rectangular(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"], ["b", "c", "d"]])
        assert read_sheet(path, pad=True) == [["a", "", ""], ["b", "c", "d"]]


class TestOrientation:
    def test_columns_orientation_transposes(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["h1", "h2"], ["v1", "v2"]], orientation="columns")
        assert read_sheet(path) == [["h1", "v1"], ["h2", "v2"]]

    def test_invalid_orientation_raises(self, tmp_path):
        with pytest.raises(ValueError):
            write_sheet(tmp_path / "book.xlsx", [["a"]], orientation="sideways")


class TestAppendRows:
    def test_appends_after_existing_rows(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"], ["b"]])
        append_rows(path, [["c"], ["d"]])
        assert read_sheet(path) == [["a"], ["b"], ["c"], ["d"]]

    def test_empty_rows_is_a_noop(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]])
        append_rows(path, [])
        assert read_sheet(path) == [["a"]]

    def test_creates_file_when_absent(self, tmp_path):
        path = tmp_path / "fresh.xlsx"
        append_rows(path, [["a"], ["b"]])
        assert read_sheet(path) == [["a"], ["b"]]


class TestDimensionGuard:
    def test_too_many_columns_raises(self, tmp_path):
        from pyhandlexl.errors import DimensionError

        path = tmp_path / "book.xlsx"
        with pytest.raises(DimensionError):
            write_sheet(path, [[""] * 16_385])


class TestSheetManagement:
    def test_list_sheets(self, tmp_path):
        path = tmp_path / "book.xlsx"
        wb = Workbook()
        wb.active.title = "One"
        wb.create_sheet("Two")
        wb.save(path)
        assert list_sheets(path) == ["One", "Two"]

    def test_list_sheets_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            list_sheets(tmp_path / "nope.xlsx")

    def test_sheet_exists(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]], sheet="Data")
        assert sheet_exists(path, "Data") is True
        assert sheet_exists(path, "Missing") is False

    def test_create_sheet_on_new_file(self, tmp_path):
        path = tmp_path / "fresh.xlsx"
        create_sheet(path, "Sales")
        assert list_sheets(path) == ["Sales"]

    def test_create_sheet_appends_to_existing(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]])  # default sheet "Sheet"
        create_sheet(path, "Extra")
        assert list_sheets(path) == ["Sheet", "Extra"]

    def test_create_sheet_duplicate_raises(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]], sheet="Data")
        with pytest.raises(ValueError):
            create_sheet(path, "Data")

    def test_create_sheet_invalid_name_raises(self, tmp_path):
        with pytest.raises(SheetNameError):
            create_sheet(tmp_path / "book.xlsx", "bad/name")

    def test_delete_sheet(self, tmp_path):
        path = tmp_path / "book.xlsx"
        wb = Workbook()
        wb.create_sheet("Gone")
        wb.save(path)
        delete_sheet(path, "Gone")
        assert sheet_exists(path, "Gone") is False

    def test_delete_missing_sheet_raises(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]])
        with pytest.raises(SheetNotFoundError):
            delete_sheet(path, "Ghost")

    def test_delete_last_sheet_raises(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]])
        with pytest.raises(ValueError):
            delete_sheet(path, list_sheets(path)[0])

    def test_rename_sheet(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]], sheet="Old")
        rename_sheet(path, "Old", "New")
        assert list_sheets(path) == ["New"]
        assert read_sheet(path, "New") == [["a"]]

    def test_rename_missing_sheet_raises(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]])
        with pytest.raises(SheetNotFoundError):
            rename_sheet(path, "Nope", "New")

    def test_rename_to_existing_name_raises(self, tmp_path):
        path = tmp_path / "book.xlsx"
        wb = Workbook()
        wb.active.title = "A"
        wb.create_sheet("B")
        wb.save(path)
        with pytest.raises(ValueError):
            rename_sheet(path, "A", "B")

    def test_rename_invalid_name_raises(self, tmp_path):
        path = tmp_path / "book.xlsx"
        write_sheet(path, [["a"]], sheet="Old")
        with pytest.raises(SheetNameError):
            rename_sheet(path, "Old", "x" * 32)
