"""Resilient workbook I/O: retry-on-lock loading and atomic saving.

Internal module — not part of the public API.
"""

from __future__ import annotations

import os
from collections.abc import Callable
from pathlib import Path
from secrets import token_hex
from time import sleep
from typing import TypeVar
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook

from pyhandlexl.errors import FileLockedError, InvalidFileError
from pyhandlexl.validate import is_valid_xlsx

_T = TypeVar("_T")

DEFAULT_RETRIES = 5
DEFAULT_DELAY = 0.5

# Raised when a file is briefly unavailable: open in Excel, or a competing
# writer has not finished flushing it yet. Worth retrying.
_TRANSIENT_ERRORS = (PermissionError, BadZipFile, EOFError)


def _retry(
    operation: Callable[[], _T],
    *,
    path: Path,
    retries: int = DEFAULT_RETRIES,
    delay: float = DEFAULT_DELAY,
) -> _T:
    """Run *operation*, retrying on transient file errors.

    Raises FileLockedError if it never succeeds.
    """
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            return operation()
        except _TRANSIENT_ERRORS as error:
            last_error = error
            if attempt < retries:
                sleep(delay)
    raise FileLockedError(
        f"could not access {path} after {retries} attempt(s): {last_error}"
    ) from last_error


def safe_load(
    path: str | Path,
    *,
    retries: int = DEFAULT_RETRIES,
    delay: float = DEFAULT_DELAY,
) -> Workbook:
    """Load an existing workbook, retrying while the file is locked.

    Raises:
        FileNotFoundError: no file at *path*.
        InvalidFileError: the file exists but is not a readable .xlsx.
        FileLockedError: the file stayed locked through every retry.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"no file at {path}")
    if not is_valid_xlsx(path):
        raise InvalidFileError(f"{path} is not a readable .xlsx workbook")
    return _retry(lambda: load_workbook(path), path=path, retries=retries, delay=delay)


def load_or_create(
    path: str | Path,
    *,
    retries: int = DEFAULT_RETRIES,
    delay: float = DEFAULT_DELAY,
) -> Workbook:
    """Load the workbook at *path*, or return a new empty one if it does not exist."""
    path = Path(path)
    if not path.exists():
        return Workbook()
    return safe_load(path, retries=retries, delay=delay)


def atomic_save(
    workbook: Workbook,
    path: str | Path,
    *,
    retries: int = DEFAULT_RETRIES,
    delay: float = DEFAULT_DELAY,
) -> None:
    """Save *workbook* to *path* without risking the file already there.

    Writes to a temporary file in the same directory, verifies it is a
    readable .xlsx, then atomically replaces the target. On any failure the
    temporary file is removed and the original is left untouched.

    Raises:
        InvalidFileError: the freshly written file failed validation.
        FileLockedError: the target stayed locked through every retry.
    """
    path = Path(path)
    tmp = path.parent / f".{path.stem}.{token_hex(6)}.tmp.xlsx"
    try:
        _retry(lambda: workbook.save(tmp), path=path, retries=retries, delay=delay)
        if not is_valid_xlsx(tmp):
            raise InvalidFileError(
                f"wrote a temporary file for {path} but it failed validation; {path} left unchanged"
            )
        _retry(lambda: os.replace(tmp, path), path=path, retries=retries, delay=delay)
    finally:
        tmp.unlink(missing_ok=True)
